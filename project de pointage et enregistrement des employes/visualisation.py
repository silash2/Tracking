from tkinter import*
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
import openpyxl, xlrd
from openpyxl import Workbook
import os
from tkinter.ttk import Combobox
import pathlib
import time
from tkinter.ttk import Combobox, Treeview, Scrollbar
import subprocess
from datetime import datetime

background = "#06283D"

   
def load_data():
    wb = openpyxl.load_workbook('Employee_data.xlsx')
    sheet = wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tree.insert('', END, values=row)
    
def Enregistre():
    root.destroy()
    subprocess.run(["python","enregistrement.py"])

def recherche_dans_base():
    
    search_info_entry= int(entry_matricule.get())
    
    if not search_info_entry:
        messagebox.showwarning("Erreur", "Veuillez entrer une matricule.")
        return

    wb= openpyxl.load_workbook('Employee_data.xlsx')
    sheet = wb.active

    # Rechercher la ligne correspondant à la matricule
    found = False
    for row in sheet.iter_rows(min_row=2):
        if row[1].value == search_info_entry:
            found = True
            row_num = row[0].row
        
            rech1 = sheet.cell(row=row_num,column=1).value
            rech2 = sheet.cell(row=row_num, column=2).value
            rech3 = sheet.cell(row=row_num, column=3).value
            rech4 = sheet.cell(row=row_num, column=4).value
            rech5 = sheet.cell(row=row_num, column=5).value
            rech6 = sheet.cell(row=row_num, column=6).value
            rech7 = sheet.cell(row=row_num, column=7).value
            rech8 = sheet.cell(row=row_num, column=8).value
            rech9 = sheet.cell(row=row_num, column=9).value
            rech10 = sheet.cell(row=row_num, column=10).value
            rech11 = sheet.cell(row=row_num, column=11).value
            rech12 = sheet.cell(row=row_num, column=12).value
            rech13 = sheet.cell(row=row_num, column=13).value
            rech14 = sheet.cell(row=row_num, column=14).value
            rech15 = sheet.cell(row=row_num, column=15).value
            
            

            tree.delete(*tree.get_children())
            tree.insert('', 'end', values=(rech1,rech2,rech3,rech4,rech5,rech6,rech7,rech8,rech9,rech10,rech11,rech12,rech13,rech14,rech15))
            break

        if search_info_entry is None:
            messagebox.showwarning("Erreur", "Matricule non trouvé.")
            return

# calcul d'ancienete   






#la fenetre de l'application
root = Tk()
root.title("List of worker")
root.geometry("1250x700+210+100")
root.config(bg=background)
root.resizable(width= False, height=False)

#entete de l'application
Label(root, text="Liste des employes",width=10,height=3,bg="white",fg='#f0687c',font="arial 20 bold").pack(side= TOP, fill=X)
Label(root, text="E-mail: wzafitsara@gmail.com", width=10,height=2, anchor='e').pack(side=BOTTOM, fill=X)

#search bar
entry_matricule = StringVar()
Entry(root, textvariable=entry_matricule, width=15, bd=2, font='arial 15 bold').place(x=920,y=40)

Srch =Button(root, text="Search", compound=LEFT, width=10,bd=2,bg='#68ddfa', font='arial 12 bold',command=recherche_dans_base).place(x=1100, y=38)

#update
update_btn =Button(root, text="Update",width=15,bd=2,font="arial 12 bold",bg="green").place(x=110, y=40)


#Date and time
Date = StringVar()
Time = StringVar()

today= date.today()

d1 = today.strftime("%d/%m/%y")
t1= time.strftime("%H:%M:%S")
Label(root, text="Date: ",font="arial 13 bold", width=10).place(x=30, y=150)
Date_entry = Entry(root, textvariable=Date, width= 13, font="arial 13 ",state='r').place(x=150, y=150)
Label(root, text="Time: ",font="arial 13 bold", width=10).place(x=300, y=150)
Time_entry = Entry(root,textvariable=Time,width=13, font= "arial 13" ,state='r').place(x=450, y=150)
Date.set(d1)
Time.set(t1)


# Treeview for displaying data
tree_frame = Frame(root)
tree_frame.place(x=40, y=200, width=1200, height=380)

scrollbary = Scrollbar(tree_frame)
scrollbary.pack(side=RIGHT, fill=Y)

scrollbarx = Scrollbar(tree_frame, orient=HORIZONTAL)
scrollbarx.pack(side=BOTTOM, fill=X)

tree = Treeview(tree_frame, columns=("index", "matricule", "nom", "prenom", "categorie", "fonction", "adresse", "salaire", "cin", "date_entree", "ancienete", "debauche", "telephone", "genre", "date_naissance"), show='headings', yscrollcommand=scrollbary.set, xscrollcommand= scrollbarx.set)

scrollbary.config(command=tree.yview)
scrollbarx.config(command=tree.xview)

tree.heading("index", text="Index")
tree.heading("matricule", text="Matricule")
tree.heading("nom", text="Nom")
tree.heading("prenom", text="Prenom")
tree.heading("categorie", text="Categorie")
tree.heading("fonction", text="Fonction")
tree.heading("adresse", text="Adresse")
tree.heading("salaire", text="Salaire")
tree.heading("cin", text="CIN")
tree.heading("date_entree", text="Date Entree")
tree.heading("ancienete", text="Ancienete")
tree.heading("debauche", text="Debauche")
tree.heading("telephone", text="Telephone")
tree.heading("genre", text="Genre")
tree.heading("date_naissance", text="Date Naissance")

tree.pack(fill=BOTH, expand=1)

# Load data into Treeview
load_data()

Button(root, text="Nouveau", width=15, height=3, bg="lightgreen", command=Enregistre).place(x=35,y=600)

root.mainloop()