from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import openpyxl
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import os
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from tkinter.ttk import Combobox, Treeview, Scrollbar
import subprocess, math


background = "#06283D"

month_year = datetime.now().strftime("%Y-%m")
directory = f"data/{month_year}/"
filename = f"{directory}Enregistrements_{month_year}.xlsx"

# Créer les dossiers si nécessaire
if not os.path.exists(directory):
    os.makedirs(directory)
if not os.path.exists(f"fiches_de_paie/{month_year}"):
    os.makedirs(f"fiches_de_paie/{month_year}")

def get_workbook_for_month(): 
    month_year = datetime.now().strftime("%Y-%m")
    filename = f"{directory}Enregistrements_{month_year}.xlsx"
    if not os.path.exists(filename):
        
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Enregistrements"

        # Créer les en-têtes
        headers = ["Matricule", "Nom","prenom","Salaire base","categories", "temps de presence", "salaire brut","CNaPS","OSTIE","salaire imposable", "IRSA", "avance","retenue","reste avance", "salaire_net"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

        #completer les donnees de employee  
        get_data_base()
        for row_number, (matricule, (nom, prenom, salaire_base, categories)) in enumerate(matricules.items(), start=2):
            sheet.cell(row=row_number, column=1, value=matricule)
            sheet.cell(row=row_number, column=2, value=nom)
            sheet.cell(row=row_number,column=3,value=prenom)
            sheet.cell(row=row_number,column=4, value=salaire_base)
            sheet.cell(row= row_number, column=5, value=categories)
        workbook.save(filename)
    else:
        workbook = load_workbook(filename)
        
    return workbook, filename


#recuperer des donnees de employee_data.xlsx
def get_data_base():
    global base
    file= openpyxl.load_workbook('Employee_data.xlsx')
    base = file.active
    global matricules 
    matricules = {}

    for row in base.iter_rows(min_row=2):
        matricule = row[1].value
        nom = row[2].value
        prenom = row[3].value
        salaire_base = row[7].value
        categories = row[4].value
        matricules[matricule]=(nom, prenom, salaire_base, categories)
            
    return matricules

# Charger les données dans le Treeview
def load_data():
    # Obtenir le workbook et la feuille pour aujourd'hui
    workbook, filename = get_workbook_for_month()
    today_str = datetime.now().strftime("%d-%m-%Y")

    if today_str not in workbook.sheetnames:
        messagebox.showinfo("Info", "Aucune donnée disponible pour aujourd'hui.")
        return

    sheet = workbook[today_str]
    
    # Effacer les anciennes données dans le Treeview
    tree.delete(*tree.get_children())

    # Charger les nouvelles données pour aujourd'hui
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tree.insert('', 'end', values=row)

    workbook.close()

#enregistre l'heure de travaille et calcul les salaires des employes
def enregistrer_heure(event_type):
    now = datetime.now()
    date_str = now.strftime("%d/%m/%Y")
    date_stt = now.strftime("%d-%m-%Y")
    time_str = now.strftime("%H:%M:%S")
    
    workbook, filename = get_workbook_for_month()
    
    #donne la date de l'enregistrement a la feuille 
    if date_stt not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=date_stt)
        sheet.append(["Matricule", "Nom", "Prénom", "Salaire Base","categories", "Date", "Entrée Matin", "Sortie Matin", "Entrée Après-midi", "Sortie Après-midi", "Heures Travaillées", "Heures Supplémentaires", "Salaire"])
    else:
        sheet = workbook[date_stt]
    
    get_data_base()  # Mettre à jour les données des employés
    matricules = get_data_base()
    for row_number, (matricule, (nom, prenom, salaire_base,categories)) in enumerate(matricules.items(), start=2):
        sheet.cell(row=row_number, column=1, value=matricule)
        sheet.cell(row=row_number, column=2, value=nom)
        sheet.cell(row=row_number, column=3, value=prenom)
        sheet.cell(row=row_number, column=4, value=salaire_base)
        sheet.cell(row=row_number, column=5, value=categories)
        sheet.cell(row=row_number, column=6, value=date_str)
    workbook.save(filename)

    for row_number, matricule in enumerate(matricules.keys(), start=2):
        if event_type == "entrée_matin":
            if sheet.cell(row=row_number, column=7, value=time_str).value!=" ":
                sheet.cell(row=row_number, column=7, value=time_str)
            else:
                pass
            
        elif event_type == "sortie_matin":
             
            if sheet.cell(row=row_number, column=8, value=time_str).value!= " ":
                    sheet.cell(row=row_number, column=8, value=time_str)
            else:
                pass
        
        elif event_type == "entrée_apres_midi":
            if sheet.cell(row=row_number, column=9, value=time_str).value!=" ":
                    sheet.cell(row=row_number, column=9, value=time_str)
            else:
                pass
        elif event_type == "sortie_après_midi":
            if not sheet.cell(row=row_number, column=10, value=time_str).value:
                    sheet.cell(row=row_number, column=10, value=time_str)
            else:
                pass

    workbook.save(filename)

    messagebox.showinfo("Succès", f"Heure d'{event_type.replace('_', ' ')} enregistrée pour {date_str} à {time_str}.")
    tree.delete(*tree.get_children())
    load_data()

#calcul salaire pour la personne en question
def calcul_salaire():
    now = datetime.now()
    date_str = now.strftime("%d/%m/%Y")
    date_stt = now.strftime("%d-%m-%Y")
    time_str = now.strftime("%H:%M:%S")
    
    workbook, filename = get_workbook_for_month()
    
    #donne la date de l'enregistrement a la feuille 
    if date_stt not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=date_stt)
        sheet.append(["Matricule", "Nom", "Prénom", "Salaire Base","categories", "Date", "Entrée Matin", "Sortie Matin", "Entrée Après-midi", "Sortie Après-midi", "Heures Travaillées", "Heures Supplémentaires", "Salaire"])
    else:
        sheet = workbook[date_stt]
    
    get_data_base()  # Mettre à jour les données des employés
    matricules = get_data_base()
    for row_number, (matricule, (nom, prenom, salaire_base,categories)) in enumerate(matricules.items(), start=2):
        sheet.cell(row=row_number, column=1, value=matricule)
        sheet.cell(row=row_number, column=2, value=nom)
        sheet.cell(row=row_number, column=3, value=prenom)
        sheet.cell(row=row_number, column=4, value=salaire_base)
        sheet.cell(row=row_number, column=5, value=categories)
        sheet.cell(row=row_number, column=6, value=date_str)
    workbook.save(filename)
    if not (sheet.cell(row=row_number, column=7).value and
        sheet.cell(row=row_number, column=8).value and
        sheet.cell(row=row_number, column=9).value and
        sheet.cell(row=row_number, column=10).value):
        pass
    else:
        for row_number, (matricule, (nom, prenom, salaire_base, categories)) in enumerate(matricules.items(), start=2):
            # Calcul des heures seulement si c'est nécessaire
            heures_pause = 30  # Pause de 30 minutes
            entree_matin = datetime.strptime(sheet.cell(row=row_number, column=7).value, "%H:%M:%S")
            sortie_matin = datetime.strptime(sheet.cell(row=row_number, column=8).value, "%H:%M:%S")
            entree_apres_midi = datetime.strptime(sheet.cell(row=row_number, column=9).value, "%H:%M:%S")
            sortie_apres_midi = datetime.strptime(sheet.cell(row=row_number, column=10).value, "%H:%M:%S")

            heures_matin = (sortie_matin - entree_matin).seconds/60
            heures_apres_midi = (sortie_apres_midi - entree_apres_midi).seconds/60

            heures_travaillees = heures_matin + heures_apres_midi 


            # Conversion des heures travaillées en format HH:MM
            total_minutes = int(heures_travaillees)
            hours = total_minutes // 60
            minutes = total_minutes % 60
            heures_travaillees_str = f"{hours:02}:{minutes:02}"
            sheet.cell(row=row_number, column=11, value=heures_travaillees_str)

            # Calcul des heures normales et des heures supplémentaires en minutes
            heures_normales = 8 * 60  # 8 heures en minutes

            #verifier la categories
            if sheet.cell(row=row_number, column=5, value=categories).value!= "HC":
                heures_sup = max(0, heures_travaillees - heures_normales)
                # Conversion des heures supplémentaires en format HH:MM
                total_minutes_sup = int(heures_sup)
                hours_sup = total_minutes_sup // 60
                minutes_sup = total_minutes_sup % 60
                heures_sup_str = f"{hours_sup:02}:{minutes_sup:02}"
                sheet.cell(row=row_number, column=12, value=heures_sup_str)
            else:
                heures_sup = "00:00"
                heures_sup_str = heures_sup
                sheet.cell(row=row_number, column=12, value=heures_sup_str)
            

            # Calcul du salaire
            total_minutes_normales = round(173.33 * 60 ) # 173.33 heures en minutes
            salaire_par_minute = math.ceil((salaire_base / total_minutes_normales))
            salaire = math.ceil((heures_travaillees * salaire_par_minute))
            sheet.cell(row=row_number, column=13, value=salaire)
        
    workbook.save(filename)


#addition les salaires recus par jour par feuille autres que enregistrements
def sum_salaire():
    workbook, filename = get_workbook_for_month()
    enregistrements_sheet = workbook["Enregistrements"]
    matricules = get_data_base()
    total_salaires = {matricule: 0 for matricule in matricules}
    total_heures_presence = {matricule: 0 for matricule in matricules}

    for sheet_name in workbook.sheetnames:
        if sheet_name != "Enregistrements":
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                matricule = row[0]
                salaire = row[12] if row[12] else 0
                heures_travaillees_str = row[10] if row[10] else "00:00"
                heures_travaillees = int(heures_travaillees_str.split(':')[0]) * 60 + int(heures_travaillees_str.split(':')[1])
                if matricule in total_salaires:
                    total_salaires[matricule] += salaire
                    total_heures_presence[matricule] += heures_travaillees
                else:
                    total_salaires[matricule]=salaire
                    total_heures_presence[matricule]=heures_travaillees
                
    for row_number, (matricule, (nom, prenom, salaire_base,categories)) in enumerate(matricules.items(), start=2):
        salaire_brut = math.ceil((total_salaires[matricule])/500)*500
        temps_presence = total_heures_presence[matricule]
        heures = temps_presence // 60
        minutes = temps_presence % 60
        temps_presence_str = f"{heures:02}:{minutes:02}"


        if math.ceil(salaire_brut*0.01)<3000:
            cnaps = 3000
        else:

            cnaps = math.ceil(salaire_brut*0.01/100)*100

        salaire_imp = math.ceil((salaire_brut-cnaps)/500)*500
        

        #calcul irsa
        if salaire_imp != 0:
            part1 = max(0, salaire_imp - 350000)
            part2 = min(part1, 50000) * 0.05
                
            part3 = max(0, salaire_imp - 400000)
            part4 = min(part3, 100000) * 0.10
                
            part5 = max(0, salaire_imp - 500000)
            part6 = min(part5, 100000) * 0.15
                
            part7 = max(0, salaire_imp - 600000) * 0.20

            irsa = math.ceil(max(3000, 0 + part2 + part4 + part6 + part7)/100)*100
        else:
            irsa=3000
        salaire_net = math.ceil((salaire_brut - cnaps - irsa)/500)*500

        enregistrements_sheet.cell(row=row_number, column=5, value=categories)
        enregistrements_sheet.cell(row=row_number, column=6, value=temps_presence_str)
        enregistrements_sheet.cell(row=row_number, column=7, value=salaire_brut)
        enregistrements_sheet.cell(row=row_number, column=8, value=cnaps)
        enregistrements_sheet.cell(row=row_number, column=11,value=irsa)
        enregistrements_sheet.cell(row=row_number, column=10,value=salaire_imp)
        enregistrements_sheet.cell(row=row_number, column=15,value=salaire_net)

        

    workbook.save(filename)
    messagebox.showinfo("Succès", "Les salaires et les temps de présence ont été mis à jour dans la feuille Enregistrements.")



#search for the special employee for absence  or permission or for
def marquer_absence_conge_permissions():
    
    search_info_entry= entry_matricule.get()
    
    if not search_info_entry:
        messagebox.showwarning("Erreur", "Veuillez entrer une matricule.")
        return

    now = datetime.now()
    date_str = now.strftime("%d/%m/%Y")
    date_stt = now.strftime("%d-%m-%Y")
    time_str = now.strftime("%H:%M:%S")
    
    workbook, filename = get_workbook_for_month()

    if date_stt not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=date_stt)
        sheet.append(["Matricule", "Nom", "Prénom", "Salaire Base","categories", "Date", "Entrée Matin", "Sortie Matin", "Entrée Après-midi", "Sortie Après-midi", "Heures Travaillées", "Heures Supplémentaires", "Salaire"])
    else:
        sheet = workbook[date_stt]


    # Rechercher la ligne correspondant à la matricule
    found = False
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == int(search_info_entry):
            found = True
            row_num = row[0].row
            rech1 = sheet.cell(row=row_num,column=1).value
            rech2 = sheet.cell(row=row_num, column=2).value
            rech3 = sheet.cell(row=row_num, column=3).value
            rech4 = sheet.cell(row=row_num, column=4).value
            rech5 = sheet.cell(row=row_num, column=5).value
            rech6 = sheet.cell(row=row_num, column=6).value
            

            tree.delete(*tree.get_children())
            tree.insert('', 'end', values=(rech1,rech2,rech3,rech4,rech5,rech6))

            messagebox.showinfo("Succès", f"Statut  enregistré pour la matricule {search_info_entry}.")
            break
        else:
            messagebox.showwarning("warning", f"le matricule {search_info_entry} ne correspond a aucun resulultat")
            

        if search_info_entry is None:
            messagebox.showwarning("Erreur", "Matricule non trouvé.")
            return

        
    # Enregistrer le fichier et afficher un message de succès
    workbook.save(filename)
    tree.delete(*tree.get_children())
    load_data()

#pour marquer que tout le monde est absent ou bien en permission
def marquer_absence_conge_permission(event_type):
    workbook, filename = get_workbook_for_month()
    date_stt = datetime.now().strftime("%d-%m-%Y")
    if date_stt not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=date_stt)
        sheet.append(["Matricule", "Nom", "Prénom", "Salaire Base", "Date", "Entrée Matin", "Sortie Matin", "Entrée Après-midi", "Sortie Après-midi", "Heures Travaillées", "Heures Supplémentaires", "Salaire"])
    else:
        sheet = workbook[date_stt]
    
    get_data_base()
    for row_number, (matricule, (nom, prenom, salaire_base)) in enumerate(matricules.items(), start=2):
        if event_type == "absence":
            sheet.cell(row=row_number, column=6, value="ABS")
            sheet.cell(row=row_number, column=7, value="ABS")
            sheet.cell(row=row_number, column=8, value="ABS")
            sheet.cell(row=row_number, column=9, value="ABS")
            sheet.cell(row=row_number, column=10, value="0")
            sheet.cell(row=row_number, column=11, value="0")
            sheet.cell(row=row_number, column=12, value="0")
        elif event_type == "congé":
            sheet.cell(row=row_number, column=6, value="CONG")
            sheet.cell(row=row_number, column=7, value="CONG")
            sheet.cell(row=row_number, column=8, value="CONG")
            sheet.cell(row=row_number, column=9, value="CONG")
            sheet.cell(row=row_number, column=10, value="0")
            sheet.cell(row=row_number, column=11, value="0")
            sheet.cell(row=row_number, column=12, value="0")
        elif event_type == "permission":
            sheet.cell(row=row_number, column=6, value="PERM")
            sheet.cell(row=row_number, column=7, value="PERM")
            sheet.cell(row=row_number, column=8, value="PERM")
            sheet.cell(row=row_number, column=9, value="PERM")
            sheet.cell(row=row_number, column=10, value="0")
            sheet.cell(row=row_number, column=11, value="0")
            sheet.cell(row=row_number, column=12, value="0")
    
    workbook.save(filename)
    load_data()
    messagebox.showinfo("Succès", f"{event_type.replace('_', ' ')} marquée pour aujourd'hui.")



def imprimer_fiche_de_paie():
    workbook, _ = get_workbook_for_month()
    sheet = workbook.active
    cumul_heures = timedelta()
    total_salaire = 0.0
    heures_sup_totales = timedelta()
    CNaPS= 0.01
    IRSA= float()
    Avance = DoubleVar()

    for row in sheet.iter_rows(values_only=True):
       
            if row[6]:
                heures_travaillees = datetime.strptime(row[6], "%H:%M:%S")
                cumul_heures += timedelta(hours=heures_travaillees.hour, minutes=heures_travaillees.minute, seconds=heures_travaillees.second)
            if row[7]:
                heures_sup = datetime.strptime(row[7], "%H:%M:%S")
                heures_sup_totales += timedelta(hours=heures_sup.hour, minutes=heures_sup.minute, seconds=heures_sup.second)
            if row[8]:
                total_salaire += row[8]

    workbook.close()
    
    month_year = datetime.now().strftime("%Y-%m")
    pdf_filename = f"fiches_de_paie/Fiche_de_paie__{month_year}.pdf"
    c = canvas.Canvas(pdf_filename, pagesize=letter)
    c.drawString(100, 750, f"Fiche de paie pour")
    c.drawString(100, 730, f"Heures travaillées: {cumul_heures}")
    c.drawString(100, 710, f"Heures supplémentaires: {heures_sup_totales}")
    c.drawString(100, 710, f"Heures supplémentaires: {CNaPS}")
    c.drawString(100, 710, f"Heures supplémentaires: {IRSA}")
    c.drawString(100, 710, f"Heures supplémentaires: {Avance}")
    c.drawString(100, 690, f"Salaire total: {total_salaire} Ar")
    c.save()
    
    messagebox.showinfo("Fiche de paie", f"Fiche de paie générée: {pdf_filename}")

#fait appel a visualisation.py 
def vision():
    root.destroy()
    subprocess.run(["python","visualisation.py"])

#fait appel a enregistrement.py
def enregistrement():
    root.destroy()
    subprocess.run(["python", "enregistrement.py"])

#fait appel a pointage.py
def pointage():
    root.destroy()
    subprocess.run(["python", "pointage.py"])
# Interface graphique
root = Tk()
root.title("Pointage des heures")
root.geometry("1250x700+210+100")
root.config(bg=background)
root.resizable(width= False, height=False)

#search bar
enter_mat= StringVar()
Label(root, text="Matricules:", font="arial 13 bold", width=15).place(x=700, y=115)
entry_matricule = Entry(root,textvariable=enter_mat, width=25,bd=2, font="arial 13")
entry_matricule.place(x=845, y=115)
Srch =Button(root, text="Search", compound=LEFT, width=10,bd=2,bg='#68ddfa', font='arial 12 bold', command=marquer_absence_conge_permissions).place(x=1100, y=110)

#update
update_btn =Button(root, text="Update",width=30,  font="arial 12",bg="green", command=sum_salaire).place(x=30, y=450)

#entete de fenetre
Label(root, text="Pointages et calcul de salaires",width=10,height=3,bg="white",fg='#f0687c',font="arial 20 bold").pack(side= TOP, fill=X)

#pied de page de tous les fenetres en utilisation
Label(root, text="E-mail: wzafitsara@gmail.com", width=10,height=2, anchor='e').pack(side=BOTTOM, fill=X)


# Boutons pour enregistrer les heures
Button(root, text="Entrée matin", width=15, height=2, bg="lightgreen",  command=lambda: enregistrer_heure("entrée_matin")).place(x=30, y=200)
Button(root, text="Sortie matin", width=15, height=2, bg="lightgreen", command=lambda: enregistrer_heure("sortie_matin")).place(x=200, y=200)
Button(root, text="Entrée après-midi", width=15, height=2, bg="lightgreen", command=lambda: enregistrer_heure("entrée_apres_midi")).place(x=30, y=260)
Button(root, text="Sortie après-midi", width=15, height=2, bg="lightgreen", command=lambda: enregistrer_heure("sortie_après_midi")).place(x=200, y=260)

# Boutons pour calculer le cumul des heures et imprimer la fiche de paie
Button(root, text="Imprimer fiche de paie", width=30, bg="lightblue", font="arial 12",  command=imprimer_fiche_de_paie).place(x=30, y=400)

#button pour aller vers visualisation
Button(root, text="Listes",command=vision,width=30, font="arial 12", bg="blue").place(x=30, y=550)
Button(root, text="Nouveau", command=enregistrement, font="arial 12", width=30, bg="gray").place(x=30, y=500)
Button(root, text="Entrer les heures", command=pointage,width=30, font="arial 12", bg="yellow").place(x=30, y=550)

# Arbre pour afficher les données
tree_frame = Frame(root)
tree_frame.place(x=450, y=150, width=750, height=500)

scrollbary = Scrollbar(tree_frame)
scrollbary.pack(side=RIGHT, fill=Y)

scrollbarx = Scrollbar(tree_frame, orient=HORIZONTAL)
scrollbarx.pack(side=BOTTOM, fill=X)

tree = Treeview(tree_frame, columns=("Matricule","Nom","Prenom","Salaire base","Categories", "Date", "Entrée matin", "Sortie matin", "Entrée après-midi", "Sortie après-midi", "Heures travaillées", "Heures supplémentaires", "Salaire"), show='headings', yscrollcommand=scrollbary.set, xscrollcommand=scrollbarx.set)

scrollbary.config(command=tree.yview)
scrollbarx.config(command=tree.xview)

tree.heading("Matricule", text="Matricules")
tree.heading("Nom", text="Nom")
tree.heading("Prenom", text="Prenom")
tree.heading("Salaire base", text="Salaire de base")
tree.heading("Categories", text="Categories")
tree.heading("Date", text="Date")
tree.heading("Entrée matin", text="Entrée matin")
tree.heading("Sortie matin", text="Sortie matin")
tree.heading("Entrée après-midi", text="Entrée après-midi")
tree.heading("Sortie après-midi", text="Sortie après-midi")
tree.heading("Heures travaillées", text="Heures travaillées")
tree.heading("Heures supplémentaires", text="Heures supplémentaires")
tree.heading("Salaire", text="Salaire")

tree.pack(fill=BOTH, expand=1)

# Charger les données dans l'arbre
load_data()
root.mainloop()

