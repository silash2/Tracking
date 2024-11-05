from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import openpyxl
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
import os
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from tkinter.ttk import Combobox, Treeview, Scrollbar
import subprocess, math



background = "#06283D"

month_year = datetime.now().strftime("%Y-%m")
directory = f"data/{month_year}/"
filename = f"{directory}Enregistrements_{month_year}.xlsx"

# Créer les dossiers si nécessaire
if not os.path.exists(directory):
    os.makedirs(directory)
if not os.path.exists(f"fiches_de_paie/{month_year}"):
    os.makedirs(f"fiches_de_paie/{month_year}")

def get_workbook_for_month(): 
    month_year = datetime.now().strftime("%Y-%m")
    filename = f"{directory}Enregistrements_{month_year}.xlsx"
    if not os.path.exists(filename):
        
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Enregistrements"

        # Créer les en-têtes
        headers = ["Matricule", "Nom","prenom","Salaire base","categories", "temps de presence", "salaire brut","CNaPS","OSTIE","salaire imposable", "IRSA", "avance","retenue","reste avance", "salaire_net"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

        #completer les donnees de employee  
        get_data_base()
        for row_number, (matricule, (nom, prenom, salaire_base, categories)) in enumerate(matricules.items(), start=2):
            sheet.cell(row=row_number, column=1, value=matricule)
            sheet.cell(row=row_number, column=2, value=nom)
            sheet.cell(row=row_number,column=3,value=prenom)
            sheet.cell(row=row_number,column=4, value=salaire_base)
            sheet.cell(row= row_number, column=5, value=categories)
        workbook.save(filename)
    else:
        workbook = load_workbook(filename)
        
    return workbook, filename


#recuperer des donnees de employee_data.xlsx
def get_data_base():
    global base
    file= openpyxl.load_workbook('Employee_data.xlsx')
    base = file.active
    global matricules 
    matricules = {}

    for row in base.iter_rows(min_row=2):
        matricule = row[1].value
        nom = row[2].value
        prenom = row[3].value
        salaire_base = row[7].value
        categories = row[4].value
        matricules[matricule]=(nom, prenom, salaire_base, categories)
            
    return matricules

# Charger les données dans le Treeview
def load_data():
    # Obtenir le workbook et la feuille pour aujourd'hui
    workbook, filename = get_workbook_for_month()
    today_str = datetime.now().strftime("%d-%m-%Y")

    if today_str not in workbook.sheetnames:
        messagebox.showinfo("Info", "Aucune donnée disponible pour aujourd'hui.")
        return

    sheet = workbook[today_str]
    
    # Effacer les anciennes données dans le Treeview
    tree.delete(*tree.get_children())

    # Charger les nouvelles données pour aujourd'hui
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tree.insert('', 'end', values=row)

    workbook.close()
    
#entrer temps pour une personne 
def enter_times():
    search_info_entry = int(entry_matricule.get())
    if not search_info_entry:
        messagebox.showwarning("Erreur", "Veuillez entrer une matricule.")
        return

    date_input = entry_date.get()
    if not date_input:
        date_input = datetime.now().strftime("%d-%m-%Y")

    entree_matin = entry_entree_matin.get()
    sortie_matin = entry_sortie_matin.get()
    entree_apres_midi = entry_entree_apres_midi.get()
    sortie_apres_midi = entry_sortie_apres_midi.get()

    workbook, filename = get_workbook_for_month()

    if date_input not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=date_input)
        sheet.append(["Matricule", "Nom", "Prénom", "Salaire Base", "Catégories", "Date", "Entrée Matin", "Sortie Matin", "Entrée Après-midi", "Sortie Après-midi", "Heures Travaillées", "Heures Supplémentaires", "Salaire"])
    else:
        sheet = workbook[date_input]

    found = False
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == search_info_entry:
            found = True
            row_num = row[0].row

            if entree_matin:
                sheet.cell(row=row_num, column=7, value=entree_matin)
            if sortie_matin:
                sheet.cell(row=row_num, column=8, value=sortie_matin)
            if entree_apres_midi:
                sheet.cell(row=row_num, column=9, value=entree_apres_midi)
            if sortie_apres_midi:
                sheet.cell(row=row_num, column=10, value=sortie_apres_midi)

            messagebox.showinfo("Succès", f"Heures enregistrées pour la matricule {search_info_entry}.")
            break

    if not found:
        messagebox.showwarning("Erreur", "Matricule non trouvé.")

    workbook.save(filename)
    load_data()

#entree l'heure pour tous
def enter_heur_tous():
    entree_matin = entry_entree_matin.get()
    sortie_matin = entry_sortie_matin.get()
    entree_apres_midi = entry_entree_apres_midi.get()
    sortie_apres_midi = entry_sortie_apres_midi.get()

    now = datetime.now()
    date_str = now.strftime("%d/%m/%Y")
    date_stt = now.strftime("%d-%m-%Y")
    time_str = now.strftime("%H:%M:%S")
    
    workbook, filename = get_workbook_for_month()
    
    #donne la date de l'enregistrement a la feuille 
    if date_stt not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=date_stt)
        sheet.append(["Matricule", "Nom", "Prénom", "Salaire Base","categories", "Date", "Entrée Matin", "Sortie Matin", "Entrée Après-midi", "Sortie Après-midi", "Heures Travaillées", "Heures Supplémentaires", "Salaire"])
    else:
        sheet = workbook[date_stt]
    matricules = get_data_base()
    

    for row_number, (matricule, (nom, prenom, salaire_base,categories)) in enumerate(matricules.items(), start=2):
        sheet.cell(row=row_number, column=1, value=matricule)
        sheet.cell(row=row_number, column=2, value=nom)
        sheet.cell(row=row_number, column=3, value=prenom)
        sheet.cell(row=row_number, column=4, value=salaire_base)
        sheet.cell(row=row_number, column=5, value=categories)
        sheet.cell(row=row_number, column=6, value=date_str)
    workbook.save(filename)


    for row_number, matricule in enumerate(matricules.keys(), start=2):
        if entree_matin:
            sheet.cell(row=row_number, column=7, value=entree_matin)
        if sortie_matin:
                sheet.cell(row=row_number, column=8, value=sortie_matin)
        if entree_apres_midi:
            sheet.cell(row=row_number, column=9, value=entree_apres_midi)
        if sortie_apres_midi:
            sheet.cell(row=row_number, column=10, value=sortie_apres_midi)
           
    workbook.save(filename)

    # Effacer les anciennes données dans le Treeview
    tree.delete(*tree.get_children())

    # Charger les nouvelles données pour aujourd'hui
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tree.insert('', 'end', values=row)


    messagebox.showinfo("succes", "enregistrement de l'heure effectuer pour tous")

def calcul_salaire():
    now = datetime.now()
    date_str = now.strftime("%d/%m/%Y")
    date_stt = now.strftime("%d-%m-%Y")
    time_str = now.strftime("%H:%M:%S")
    
    workbook, filename = get_workbook_for_month()
    
    #donne la date de l'enregistrement a la feuille 
    if date_stt not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=date_stt)
        sheet.append(["Matricule", "Nom", "Prénom", "Salaire Base","categories", "Date", "Entrée Matin", "Sortie Matin", "Entrée Après-midi", "Sortie Après-midi", "Heures Travaillées", "Heures Supplémentaires", "Salaire"])
    else:
        sheet = workbook[date_stt]
    
    get_data_base()  # Mettre à jour les données des employés
    matricules = get_data_base()
    for row_number, (matricule, (nom, prenom, salaire_base,categories)) in enumerate(matricules.items(), start=2):
        sheet.cell(row=row_number, column=1, value=matricule)
        sheet.cell(row=row_number, column=2, value=nom)
        sheet.cell(row=row_number, column=3, value=prenom)
        sheet.cell(row=row_number, column=4, value=salaire_base)
        sheet.cell(row=row_number, column=5, value=categories)
        sheet.cell(row=row_number, column=6, value=date_str)
    workbook.save(filename)
    if not (sheet.cell(row=row_number, column=7).value and
        sheet.cell(row=row_number, column=8).value and
        sheet.cell(row=row_number, column=9).value and
        sheet.cell(row=row_number, column=10).value):
        pass
    else:
        for row_number, (matricule, (nom, prenom, salaire_base, categories)) in enumerate(matricules.items(), start=2):
            # Calcul des heures seulement si c'est nécessaire
            heures_pause = 30  # Pause de 30 minutes
            entree_matin = datetime.strptime(sheet.cell(row=row_number, column=7).value, "%H:%M:%S")
            sortie_matin = datetime.strptime(sheet.cell(row=row_number, column=8).value, "%H:%M:%S")
            entree_apres_midi = datetime.strptime(sheet.cell(row=row_number, column=9).value, "%H:%M:%S")
            sortie_apres_midi = datetime.strptime(sheet.cell(row=row_number, column=10).value, "%H:%M:%S")

            heures_matin = (sortie_matin - entree_matin).seconds/60
            heures_apres_midi = (sortie_apres_midi - entree_apres_midi).seconds/60

            heures_travaillees = heures_matin + heures_apres_midi 
            heures_sup = heures_travaillees - 480
            if heures_sup <0:
                heures_sup = 0
            heures_effective = max(0,heures_travaillees - heures_sup)


            # Conversion des heures travaillées en format HH:MM
            total_minutes = int(heures_travaillees)
            hours = total_minutes // 60
            minutes = total_minutes % 60
            heures_travaillees_str = f"{hours:02}:{minutes:02}"
            sheet.cell(row=row_number, column=11, value=heures_travaillees_str)

            # Calcul des heures normales et des heures supplémentaires en minutes
            heures_normales = 8 * 60  # 8 heures en minutes

            #verifier la categories
            if sheet.cell(row=row_number, column=5, value=categories).value!= "HC":
                heures_sup = max(0, heures_travaillees - heures_normales)

                # Conversion des heures supplémentaires en format HH:MM
                total_minutes_sup = int(heures_sup)
                hours_sup = total_minutes_sup // 60
                minutes_sup = total_minutes_sup % 60
                heures_sup_str = f"{hours_sup:02}:{minutes_sup:02}"
                sheet.cell(row=row_number, column=12, value=heures_sup_str)
            else:
                heures_sup = "00:00"
                heures_sup_str = heures_sup
                sheet.cell(row=row_number, column=12, value=heures_sup_str)
                
            

            # Calcul du salaire
            total_minutes_normales = round(173.33 * 60 ) # 173.33 heures en minutes
            salaire_par_minute = math.ceil((salaire_base / total_minutes_normales))
            heures_sup = int(heures_sup)
            if hours_sup <= 32 :
                    hours_sup = hours_sup*1.3
                    salaire = math.ceil((heures_effective * salaire_par_minute)+(hours_sup))
            elif hours_sup >32:
                    hours_sup_50 = hours_sup-32
                    hours_sup = (32*1.3)+(1.5*hours_sup_50)
                    salaire = math.ceil((heures_effective * salaire_par_minute)+(hours_sup))
            
            
            sheet.cell(row=row_number, column=13, value=salaire)


        tree.delete(*tree.get_children())
            # Charger les nouvelles données pour aujourd'hui
        for row in sheet.iter_rows(min_row=2, values_only=True):
            tree.insert('', 'end', values=row)
        
    workbook.save(filename)

#recherche de date ou ajoute la date si n'exste pas
def date_search():
    date_input = entry_date.get().replace("/",  "-")
    workbook, filename = get_workbook_for_month()
    get_data_base()

    if date_input not in workbook.sheetnames:
        messagebox.showinfo("info", f"il n'y a pas d'enregistrement pour {date_input}")

        sheet = workbook.create_sheet(title=date_input)
        sheet.append(["Matricule", "Nom", "Prénom", "Salaire Base","categories", "Date", "Entrée Matin", "Sortie Matin", "Entrée Après-midi", "Sortie Après-midi", "Heures Travaillées", "Heures Supplémentaires", "Salaire"])
    else:
        sheet = workbook[date_input]
    

    # Effacer les anciennes données dans le Treeview
    tree.delete(*tree.get_children())

    # Charger les nouvelles données pour aujourd'hui
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tree.insert('', 'end', values=row)


    workbook.save(filename)
    

#recherche de matricule dans base de donnee
def recherche_dans_base():
    
    search_info_entry= int(entry_matricule.get())
    
    if not search_info_entry:
        messagebox.showwarning("Erreur", "Veuillez entrer une matricule.")
        return

    wb= openpyxl.load_workbook('Employee_data.xlsx')
    sheet = wb.active

    # Rechercher la ligne correspondant à la matricule
    found = False
    for row in sheet.iter_rows(min_row=2):
        if row[1].value == search_info_entry:
            found = True
            row_num = row[0].row
        
            rech1 = sheet.cell(row=row_num,column=1).value
            rech2 = sheet.cell(row=row_num, column=2).value
            rech3 = sheet.cell(row=row_num, column=3).value
            rech4 = sheet.cell(row=row_num, column=4).value
            rech5 = sheet.cell(row=row_num, column=5).value
            rech6 = sheet.cell(row=row_num, column=6).value
            rech7 = sheet.cell(row=row_num, column=7).value
            rech8 = sheet.cell(row=row_num, column=8).value
            rech9 = sheet.cell(row=row_num, column=9).value
            rech10 = sheet.cell(row=row_num, column=10).value
            rech11 = sheet.cell(row=row_num, column=11).value
            rech12 = sheet.cell(row=row_num, column=12).value
            rech13 = sheet.cell(row=row_num, column=13).value
            rech14 = sheet.cell(row=row_num, column=14).value
            rech15 = sheet.cell(row=row_num, column=15).value
            

            tree.delete(*tree.get_children())
            tree.insert('', 'end', values=(rech1,rech2,rech3,rech4,rech5,rech6,rech7,rech8,rech9,rech10,rech11,rech12,rech13,rech14,rech15))
            break

        if search_info_entry is None:
            messagebox.showwarning("Erreur", "Matricule non trouvé.")
            return

    
#enregistre les donnees dans enregistrement sheet et fait les calculs de tous les heures de travails 
def sum_salaire():
    workbook, filename = get_workbook_for_month()
    enregistrements_sheet = workbook["Enregistrements"]
    matricules = get_data_base()
    total_salaires = {matricule: 0 for matricule in matricules}
    total_heures_presence = {matricule: 0 for matricule in matricules}

    for sheet_name in workbook.sheetnames:
        if sheet_name != "Enregistrements":
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                matricule = row[0]
                salaire = row[12] if row[12] else 0
                heures_travaillees_str = row[10] if row[10] else "00:00"
                heures_travaillees = int(heures_travaillees_str.split(':')[0]) * 60 + int(heures_travaillees_str.split(':')[1])
                if matricule in total_salaires:
                    total_salaires[matricule] += salaire
                    total_heures_presence[matricule] += heures_travaillees
                else:
                    total_salaires[matricule]=salaire
                    total_heures_presence[matricule]=heures_travaillees
                
    for row_number, (matricule, (nom, prenom, salaire_base,categories)) in enumerate(matricules.items(), start=2):
        salaire_brut = math.ceil((total_salaires[matricule])/500)*500
        temps_presence = total_heures_presence[matricule]
        heures = temps_presence // 60
        minutes = temps_presence % 60
        temps_presence_str = f"{heures:02}:{minutes:02}"


        if math.ceil(salaire_brut*0.01)<3000:
            cnaps = 3000
        else:

            cnaps = math.ceil(salaire_brut*0.01/100)*100

        salaire_imp = math.ceil((salaire_brut-cnaps)/500)*500
        

        #calcul irsa
        if salaire_imp != 0:
            part1 = max(0, salaire_imp - 350000)
            part2 = min(part1, 50000) * 0.05
                
            part3 = max(0, salaire_imp - 400000)
            part4 = min(part3, 100000) * 0.10
                
            part5 = max(0, salaire_imp - 500000)
            part6 = min(part5, 100000) * 0.15
                
            part7 = max(0, salaire_imp - 600000) * 0.20

            irsa = math.ceil(max(3000, 0 + part2 + part4 + part6 + part7)/100)*100
        else:
            irsa=3000
        salaire_net = math.ceil((salaire_brut - cnaps - irsa)/500)*500

        enregistrements_sheet.cell(row=row_number, column=5, value=categories)
        enregistrements_sheet.cell(row=row_number, column=6, value=temps_presence_str)
        enregistrements_sheet.cell(row=row_number, column=7, value=salaire_brut)
        enregistrements_sheet.cell(row=row_number, column=8, value=cnaps)
        enregistrements_sheet.cell(row=row_number, column=11,value=irsa)
        enregistrements_sheet.cell(row=row_number, column=10,value=salaire_imp)
        enregistrements_sheet.cell(row=row_number, column=15,value=salaire_net)

        

    workbook.save(filename)
    messagebox.showinfo("Succès", "Les salaires et les temps de présence ont été mis à jour dans la feuille Enregistrements.")



#fait appel a visualisation.py 
def vision():
    root.destroy()
    subprocess.run(["python","visualisation.py"])

#fait appel a enregistrement.py
def enregistrement():
    root.destroy()
    subprocess.run(["python", "enregistrement.py"])

# Fonction appelée lors de la sélection dans le combobox
def on_sheet_selected(event):

    selected_sheet = sheet_combobox.get()
    workbook, filename = get_workbook_for_month()
    sheet = workbook[selected_sheet]
    
        # Effacer les anciennes données dans le Treeview
    tree.delete(*tree.get_children())

    # Charger les nouvelles données pour aujourd'hui
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tree.insert('', 'end', values=row)

    

# Interface graphique
root = Tk()
root.title("Pointage des heures")
root.geometry("1250x700+210+100")
root.config(bg=background)
root.resizable(width= False, height=False)

entry_matricule = Entry(root, width=25)
entry_date = Entry(root, width=25)
entry_entree_matin = Entry(root, width=25)
entry_sortie_matin = Entry(root, width=25)
entry_entree_apres_midi = Entry(root, width=25)
entry_sortie_apres_midi = Entry(root, width=25)

#search bar
enter_mat= StringVar()
Label(root, text="Matricules:", font="arial 13 bold", width=15).place(x=700, y=115)
entry_matricule = Entry(root,textvariable=enter_mat, width=25,bd=2, font="arial 13")
entry_matricule.place(x=845, y=115)
Srch =Button(root, text="Search", compound=LEFT, width=10,bd=2,bg='#68ddfa', font='arial 12 bold', command=enter_times).place(x=1100, y=110)

#update
update_btn =Button(root, text="Calcul salaire",width=30,  font="arial 12",bg="green", command=calcul_salaire).place(x=80, y=600)

#entete de fenetre
Label(root, text="Pointages et calcul de salaires",width=10,height=3,bg="white",fg='#f0687c',font="arial 20 bold").pack(side= TOP, fill=X)

#pied de page de tous les fenetres en utilisation
Label(root, text="E-mail: wzafitsara@gmail.com", width=10,height=2, anchor='e').pack(side=BOTTOM, fill=X)

#entry form
dates = Label(root, text="Date (jj-mm-aaaa):",width=15)
dates.place(x=160, y=120)
date_enter= entry_date.place(x=135, y=150)
Label(root, text="Entrée AM(HH:MM:SS):").place(x=63, y=275)
entry_entree_matin.place(x=40, y=300)
Label(root, text="Sortie AM(HH:MM:SS):").place(x=255, y=275)
entry_sortie_matin.place(x=255, y=300)
Label(root, text="Entrée PM(HH:MM:SS):").place(x=63, y=335)
entry_entree_apres_midi.place(x=40, y=360)
Label(root, text="Sortie PM(HH:MM:SS):").place(x=260, y=335)
entry_sortie_apres_midi.place(x=260, y=360)

#creer un enregistrement pour une date
Button(root, text="Creer pointage pour date", bg="orange",width=30, font="Arial 12", command= date_search).place(x=80, y=200)

#button pour enregistre l'heure pour tous les salaries
Button(root, text="Enregistrer pour tous",width=30, font="arial 12", bg="light green", command=enter_heur_tous).place(x=80,y=400)
Button(root, text="Enregistrer pour matricule",width=30, font="arial 12", bg="light yellow",command=enter_times).place(x=80, y=450)

#button pour aller vers visualisation
Button(root, text="Listes",command=vision,width=30, font="arial 12", bg="blue").place(x=80, y=550)
Button(root, text="Nouveau", command=enregistrement, font="arial 12", width=30, bg="gray").place(x=80, y=500)


#combobox 
# Initialiser le combobox pour les feuilles
workbook, filename = get_workbook_for_month()
sheet_names = workbook.sheetnames
sheet_combobox = ttk.Combobox(root, values=sheet_names)
sheet_combobox.place(x=280,y=120)
sheet_combobox.bind("<<ComboboxSelected>>", on_sheet_selected)
# Arbre pour afficher les données
tree_frame = Frame(root)
tree_frame.place(x=450, y=150, width=750, height=500)

scrollbary = Scrollbar(tree_frame)
scrollbary.pack(side=RIGHT, fill=Y)

scrollbarx = Scrollbar(tree_frame, orient=HORIZONTAL)
scrollbarx.pack(side=BOTTOM, fill=X)

tree = Treeview(tree_frame, columns=("Matricule","Nom","Prenom","Salaire base","Categories", "Date", "Entrée matin", "Sortie matin", "Entrée après-midi", "Sortie après-midi", "Heures travaillées", "Heures supplémentaires", "Salaire"), show='headings', yscrollcommand=scrollbary.set, xscrollcommand=scrollbarx.set)

scrollbary.config(command=tree.yview)
scrollbarx.config(command=tree.xview)

tree.heading("Matricule", text="Matricules")
tree.heading("Nom", text="Nom")
tree.heading("Prenom", text="Prenom")
tree.heading("Salaire base", text="Salaire de base")
tree.heading("Categories", text="Categories")
tree.heading("Date", text="Date")
tree.heading("Entrée matin", text="Entrée matin")
tree.heading("Sortie matin", text="Sortie matin")
tree.heading("Entrée après-midi", text="Entrée après-midi")
tree.heading("Sortie après-midi", text="Sortie après-midi")
tree.heading("Heures travaillées", text="Heures travaillées")
tree.heading("Heures supplémentaires", text="Heures supplémentaires")
tree.heading("Salaire", text="Salaire")

tree.pack(fill=BOTH, expand=1)

# Charger les données dans l'arbre
load_data()


root.mainloop()