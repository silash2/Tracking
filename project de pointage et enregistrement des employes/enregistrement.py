from tkinter import*
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import openpyxl, xlrd
from openpyxl import Workbook
import os
import pathlib
import time
from tkinter.ttk import Combobox, Treeview, Scrollbar
import subprocess
import tkinter as tk
from tkcalendar import DateEntry
from datetime import datetime

background = "#06283D"

#display the windows 
root = Tk()
root.title("worker registration system")
root.geometry("1250x700+210+100")
root.config(bg=background)
root.resizable(width= False, height=False)

#creation de base de donne excel
file = pathlib.Path('Employee_data.xlsx')

if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1']="index"
    sheet['B1']="Matricule"
    sheet['C1']="Nom"
    sheet['D1']="prenom"
    sheet['E1']="Categorie"
    sheet['F1']="Fonction"
    sheet['G1']="adresse"
    sheet['H1']="Salaire de base"
    sheet['I1']="CIN"
    sheet['J1']="date entree"
    sheet['K1']="ancienete"
    sheet['L1']="debauche"
    sheet['M1']="telephone"
    sheet['N1']="genre"
    sheet['O1']="date de naissance"

    file.save('Employee_data.xlsx')
    
#gender selection
def selection():
    value = radio.get()
    if value == True:
        gender="Male"
        print(gender)
    else:
        gender ="Female"    
    return gender

#matriculation automatique    
def matricule():
    file= openpyxl.load_workbook('Employee_data.xlsx')
    sheet = file.active
    row= sheet.row_max
    max_row_value = sheet.cell(row= row, column=2).value

    try:
        Matricule.set(max_row_value+1)
    except:
        Matricule.set("1")

#reset value on the input
def clear():
    global img, pht1, pht
    matricule = Matricule.set('')
    noms = nom.set('')
    prenoms = prenom.set('')
    categorie = Categories.set('')
    fonctions = fonction.set('')
    adr = adresse.set('')
    sal = salaire.set('')
    cin_num = cin.set('')
    date_ent = date_entree.set('')
    tel = telephone.set('')
    gender = selection()
    date_naiss = date_naissance.set('')


#call visualisation windows and close the current in use 
def visionner():
    root.destroy()
    subprocess.run(["python","visualisation.py"])
# close the current windows
def shutdown():
    root.destroy()

#open pointage and close the current in use
def pointage():
    root.destroy()
    subprocess.run(["python","salaire.py"])

#Affichage de l'image
def showImage():
    global fileName
    global img,pht1
    fileName= filedialog.askopenfilename(initialdir=os.getcwd(),title="select image file",filetypes=(("JPG","*.jpg"),("PNG","*.png"), ("All file","*.txt")))
    img= (Image.open(fileName))
    resize_img = img.resize((190,190))
    profilePhotos = ImageTk.PhotoImage(resize_img)
    pht1.config(image=profilePhotos)
    pht1.image = profilePhotos

#calcul ancienete
def calcul_ancienete(date_debut):
    
    date_actuel = datetime.now()
    delta = date_actuel - date_debut
    jour = delta.days
    ans =  jour//365
    mois = (jour % 365)//30
    jours_restant = (jour % 30)% 30

    return ans, mois, jours_restant

# Save function
def save_data():
    matricule = Matricule.get()
    noms = nom.get()
    prenoms = prenom.get()
    categorie = Categories.get()
    fonctions = fonction.get()
    adr = adresse.get()
    sal = salaire.get()
    cin_num = cin.get()
    date_ent = date_entree.get()
    tel = telephone.get()
    gender = selection()
    date_naiss = date_naissance.get()
    

    if not (matricule and noms and prenoms and categorie and fonctions and adr and sal and cin_num and date_ent and tel and gender and date_naiss):
        messagebox.showerror("Error", "All fields are required")
        return
    
    
    date_ent_dt = datetime.strptime(date_ent, "%d/%m/%Y")  # Convertir la chaîne de date en objet datetime
    today = datetime.today()
    anciennete = today.year - date_ent_dt.year - ((today.month, today.day) < (date_ent_dt.month, date_ent_dt.day))
    
 


    wb = openpyxl.load_workbook('Employee_data.xlsx')
    sheet = wb.active
    sheet.append([sheet.max_row, matricule, noms, prenoms, categorie, fonctions, adr, sal, cin_num, date_ent, anciennete, None, tel, gender, date_naiss])
    wb.save('Employee_data.xlsx')
    messagebox.showinfo("Success", "Record added successfully")
    

    #save image
    try:
        img.save(f"images/worker/"+str(matricule)+".jpg")
    except:
        messagebox.showinfo("error", "image indisponible")
    clear()

#searching 
def search():

    text= Search.get()
    clear()
    Save_button.config(state=tk.DISABLED)
    file = openpyxl.load_workbook('Employee_data.xlsx')
    sheet = file.active
   
    for row in sheet.rows:
        matricules = row[1]
        if matricules.value == int(text):
            matricule_pos = str(matricules)[14:-1]
            matricule_num = str(matricules)[15:-1]

            #charge les donnees et l'enregistre dans un variable
            info_1 = sheet.cell(row=int(matricule_num), column=1).value
            info_2 = sheet.cell(row=int(matricule_num), column=2).value
            info_3 = sheet.cell(row=int(matricule_num), column=3).value
            info_4 = sheet.cell(row=int(matricule_num), column=4).value
            info_5 = sheet.cell(row=int(matricule_num), column=5).value
            info_6 = sheet.cell(row=int(matricule_num), column=6).value
            info_7 = sheet.cell(row=int(matricule_num), column=7).value
            info_8 = sheet.cell(row=int(matricule_num), column=8).value
            info_9 = sheet.cell(row=int(matricule_num), column=9).value
            info_10 = sheet.cell(row=int(matricule_num), column=10).value
            info_14 = sheet.cell(row=int(matricule_num), column=14).value
            info_13 = sheet.cell(row=int(matricule_num), column=13).value
            info_15 = sheet.cell(row=int(matricule_num), column=15).value

            #attribuer les valeurs de la recherche
            Matricule.set(info_2)
            nom.set(info_3)
            prenom.set(info_4)
            Categories.set(info_5)
            fonction.set(info_6)
            adresse.set(info_7)
            salaire.set(info_8)
            cin.set(info_9)
            date_entree.set(info_10)
            date_naissance.set(info_15)
            telephone.set(info_13)
            if info_14=='female':
                radio_entry1.select()
            else:
                radio_entry2.select()

            img = (f"Image.open(images/worker/"+str(info_2)+".jpg")
            resize_img = img.resize((190,190))
            profilePhotos = ImageTk.PhotoImage(resize_img)
            pht1.config(image=profilePhotos)
            pht1.image = profilePhotos




   
#entete de l'application
Label(root, text="Enregistrement des employes",width=10,height=3,bg="white",fg='#f0687c',font="arial 20 bold").pack(side= TOP, fill=X)
Label(root, text="E-mail: wzafitsara@gmail.com", width=10,height=2, anchor='e').pack(side=BOTTOM, fill=X)

#search bar
Search = StringVar()
Entry(root, textvariable=Search, width=15, bd=2, font='arial 15 bold').place(x=920,y=40)

Srch =Button(root, text="Search", compound=LEFT, width=10,bd=2,bg='#68ddfa', font='arial 12 bold',command=search).place(x=1100, y=38)

#update
update_btn =Button(root, text="Update",width=15,bd=2,font="arial 12 bold",bg="green").place(x=110, y=40)


#Date and time
Date = StringVar()
Time = StringVar()

today= date.today()

d1 = today.strftime("%d/%m/%y")
t1= time.strftime("%H:%M:%S")
Label(root, text="Date: ",font="arial 13 bold", width=10).place(x=30, y=150)
Date_entry = Entry(root, textvariable=Date, width= 13, font="arial 13 ",state='r').place(x=150, y=150)
Label(root, text="Time: ",font="arial 13 bold", width=10).place(x=300, y=150)
Time_entry = Entry(root,textvariable=Time,width=13, font= "arial 13" ,state='r')
Time_entry.place(x=450, y=150)
Date.set(d1)
Time.set(t1)

#Employee details
obj = LabelFrame(root, text="A propos", font=20,bd=2,width=850,height=450,relief=GROOVE).place(x=30, y=200)

Matricule = IntVar()
Label(obj, text="Matricule: ", font="arial 13", width=10).place(x=31, y=250)
Matricule_entry = Entry(obj, textvariable= Matricule, font="arial 13")
Matricule_entry.place(x=129, y=250)

nom = StringVar()
Label(obj, text="Nom: ", font="arial 13", width=10).place(x=31, y=300)
nom_entry = Entry(obj, textvariable= nom, font="arial 13")
nom_entry.place(x=129, y=300)

prenom = StringVar()
Label(obj, text="Prenom: ", font="arial 13", width=10).place(x=31, y=350)
prenom_entry = Entry(obj, textvariable= prenom, font="arial 13")
prenom_entry.place(x=129, y=350)


Label(obj, text="Categories: ", font="arial 13", width=10).place(x=31, y=400)
Categories = Combobox(obj, values= ['M1','M2','Os1','Os2','Os3','Op1','Op2','Op3','HC1','HC'],font="arial 13",state='r')
Categories.place(x=129, y=400)

fonction = StringVar()
Label(obj, text="Fonction: ", font="arial 13", width=10).place(x=31, y=450)
fonction_entry = Entry(obj, textvariable= fonction, font="arial 13")
fonction_entry.place(x=129, y=450)


adresse = StringVar()
Label(obj, text="Adresse: ", font="arial 13", width=10).place(x=400, y=250)
adresse_entry = Entry(obj, textvariable= adresse, font="arial 13")
adresse_entry.place(x=500, y=250)

salaire = IntVar()
Label(obj, text="Salaire de base : ", font="arial 13", width=15).place(x=350, y=300)
salaire_entry = Entry(obj, textvariable= salaire, font="arial 13")
salaire_entry.place(x=500, y=300)

cin = StringVar()
Label(obj, text="Numero de CIN : ", font="arial 13", width=15).place(x=350, y=350)
cin_entry = Entry(obj, textvariable= cin, font="arial 13")
cin_entry.place(x=500, y=350)

date_entree = StringVar()
Label(obj, text="Date d'entree : ", font="arial 13", width=17).place(x=350, y=400)
date_entry = DateEntry(obj, textvariable= date_entree, date_pattern ='d/m/y', font="arial 13",)
date_entry.place(x=500, y=400)
date_entree.set(d1)

telephone = StringVar()
Label(obj, text="Numero de telephone : ", font="arial 10", width=16).place(x=350, y=450)
telephone_entry = Entry(obj, textvariable= telephone, font="arial 13")
telephone_entry.place(x=500, y=450)

radio= BooleanVar()
Label(obj, text="Sex : ", font="arial 13", width=15).place(x=31, y=500)
radio_entry1 = Radiobutton(obj, text= "Male",variable=radio,value=True, font="arial 13",command=selection)
radio_entry1.place(x=129, y=500)
radio_entry2 = Radiobutton(obj, text= "Femelle",variable = radio,value=False, font="arial 13",command=selection)
radio_entry2.place(x=229, y=500)
radio.set(value=True)


date_naissance = StringVar()
Label(obj, text="Date de naissance : ", font="arial 13", width=16).place(x=350, y=500)
date_naissance_entry = DateEntry(obj, textvariable= date_naissance,date_pattern = 'dd/mm/yy', font="arial 13")
date_naissance_entry.place(x=500, y=500)

date_naissance.set(d1)


Save_button=tk.Button(root, text="Save", width=15, bd=2, font="arial 12 bold", bg="blue", command=save_data)
Save_button.place(x=1050, y=600)


voir = Button(obj,text="Listes ", width=20, bd=2, font="arial 12 bold", bg="purple", command=visionner )
voir.place(x=51, y=600)


#images displaying on the windows 

phts= Frame(root,bd=3,bg="gray", width=200, height=200,relief=GROOVE)
phts.place(x=1000, y=200)

img= PhotoImage(file="images/search.png")
pht1 = Label(phts,bg="light gray",image=img,justify="center")
pht1.place(x=0, y=0)

#download photos boutton
upload_button=Button(root, text="Telecharger photos",width=20, height=3,font="arial 10", bg="light blue", command=showImage)
upload_button.place(x=1019, y=430)

#reset boutton
reset_button=Button(root, text="reset", width=15, bd=2, font="arial 12 bold", bg="green", command= clear)
reset_button.place(x=350, y=600)

#aller vers pointage
pointage_button=Button(root, text="pointage", width=15, bd=2, font="arial 12 bold", bg="light gray",command= pointage)
pointage_button.place(x=650, y=600)

#exit boutton
shutdown_button=Button(root, text="Exit", width=10, bd=2, font="arial 12 bold", bg="pink", command= shutdown)
shutdown_button.place(x=900, y=600)


root.mainloop()